from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import StreamingResponse
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select
from io import BytesIO
import uuid
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from app.db.session import get_session
from app.db.models import Calculation, Project
from app.api.deps import get_current_user
from app.db.models import User
from app.core.security import sanitize_filename

router = APIRouter()

# Colores
COLOR_HEADER_BG = "1C2128"
COLOR_HEADER_FG = "F0B429"
COLOR_SECTION_BG = "21262D"
COLOR_SECTION_FG = "8B949E"
COLOR_OK = "3FB950"
COLOR_FAIL = "F85149"
COLOR_VALUE = "E6EDF3"


def _style_header(cell, bold=True):
    cell.font = Font(name="Calibri", bold=bold, color=COLOR_HEADER_FG, size=11)
    cell.fill = PatternFill("solid", fgColor=COLOR_HEADER_BG)
    cell.alignment = Alignment(horizontal="left", vertical="center")


def _style_section(cell):
    cell.font = Font(name="Calibri", bold=True, color=COLOR_SECTION_FG, size=9)
    cell.fill = PatternFill("solid", fgColor=COLOR_SECTION_BG)


def _style_row(cell, value_cell=False):
    cell.font = Font(name="Calibri", color="8B949E" if not value_cell else COLOR_VALUE, size=10)
    cell.alignment = Alignment(horizontal="left")


def _add_row(ws, row: int, label: str, value, unit: str = ""):
    ws.cell(row=row, column=1, value=label)
    ws.cell(row=row, column=2, value=value)
    ws.cell(row=row, column=3, value=unit)
    _style_row(ws.cell(row=row, column=1))
    _style_row(ws.cell(row=row, column=2), value_cell=True)
    _style_row(ws.cell(row=row, column=3))


def _add_section(ws, row: int, title: str):
    cell = ws.cell(row=row, column=1, value=title)
    ws.merge_cells(f"A{row}:C{row}")
    _style_section(cell)
    return row + 1


@router.get("/{calculation_id}/xlsx")
async def export_xlsx(
    calculation_id: str,
    db: AsyncSession = Depends(get_session),
    current_user: User = Depends(get_current_user),
):
    result = await db.execute(
        select(Calculation).where(Calculation.id == uuid.UUID(calculation_id))
    )
    calc = result.scalar_one_or_none()
    if not calc:
        raise HTTPException(status_code=404, detail="Cálculo no encontrado")

    proj_result = await db.execute(select(Project).where(Project.id == calc.project_id))
    project = proj_result.scalar_one_or_none()
    if not project or str(project.owner_id) != str(current_user.id):
        raise HTTPException(status_code=403, detail="Sin acceso")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Memoria de Cálculo"

    # Anchos de columna
    ws.column_dimensions["A"].width = 36
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 14

    r = 1

    # Título
    ws.cell(row=r, column=1, value="MEMORIA DE CÁLCULO RIC")
    ws.merge_cells(f"A{r}:C{r}")
    _style_header(ws.cell(row=r, column=1))
    ws.row_dimensions[r].height = 22
    r += 1

    ws.cell(row=r, column=1, value="RIC Conductor.calc")
    ws.merge_cells(f"A{r}:C{r}")
    ws.cell(row=r, column=1).font = Font(name="Calibri", color="8B949E", size=9, italic=True)
    r += 2

    # Info del proyecto
    r = _add_section(ws, r, "PROYECTO")
    _add_row(ws, r, "Proyecto", project.name); r += 1
    _add_row(ws, r, "Cálculo", calc.name or "Sin nombre"); r += 1
    _add_row(ws, r, "Fecha", calc.created_at.strftime("%d/%m/%Y %H:%M")); r += 1
    r += 1

    # Datos de entrada
    inp = calc.input_data
    r = _add_section(ws, r, "DATOS DE ENTRADA")
    _add_row(ws, r, "Sistema", inp.get("sistema", "").capitalize()); r += 1
    _add_row(ws, r, "Tensión nominal", inp.get("tension_v"), "V"); r += 1
    _add_row(ws, r, "Potencia", inp.get("potencia_kw"), "kW"); r += 1
    _add_row(ws, r, "Factor de potencia", inp.get("factor_potencia")); r += 1
    _add_row(ws, r, "Factor de demanda", inp.get("factor_demanda")); r += 1
    _add_row(ws, r, "Longitud del circuito", inp.get("longitud_m"), "m"); r += 1
    _add_row(ws, r, "Material conductor", str(inp.get("material", "")).upper()); r += 1
    _add_row(ws, r, "Tipo de canalización", inp.get("tipo_canalizacion", "").replace("_", " ")); r += 1
    _add_row(ws, r, "Temperatura ambiente", inp.get("temp_ambiente_c"), "°C"); r += 1
    _add_row(ws, r, "Circuitos agrupados", inp.get("circuitos_agrupados")); r += 1
    _add_row(ws, r, "Altitud", inp.get("msnm"), "msnm"); r += 1
    r += 1

    # Resultados
    res = calc.result_data
    r = _add_section(ws, r, "RESULTADOS")
    _add_row(ws, r, "Sección seleccionada", calc.seccion_mm2, "mm²"); r += 1
    _add_row(ws, r, "Calibre AWG", res.get("calibre_awg")); r += 1
    _add_row(ws, r, "Corriente de diseño", res.get("i_diseno_a"), "A"); r += 1
    _add_row(ws, r, "Corriente máx. corregida", res.get("i_max_corregida_a"), "A"); r += 1
    _add_row(ws, r, "Caída de tensión", res.get("caida_pct"), "%"); r += 1
    _add_row(ws, r, "Límite caída de tensión", res.get("limite_caida_pct"), "%"); r += 1
    _add_row(ws, r, "Sección neutro", res.get("sec_neutro_mm2"), "mm²"); r += 1
    _add_row(ws, r, "Sección tierra (PE)", res.get("sec_tierra_mm2"), "mm²"); r += 1
    r += 1

    # Factores de corrección
    r = _add_section(ws, r, "FACTORES DE CORRECCIÓN")
    _add_row(ws, r, "Ft (temperatura)", res.get("ft")); r += 1
    _add_row(ws, r, "Fg (agrupamiento)", res.get("fg")); r += 1
    _add_row(ws, r, "Fa (altitud)", res.get("fa")); r += 1
    _add_row(ws, r, "Factor total", res.get("factor_total")); r += 1
    r += 1

    # Cumplimiento
    r = _add_section(ws, r, "CUMPLIMIENTO RIC")
    cumple_cell = ws.cell(row=r, column=2, value="CUMPLE" if calc.cumple_ric else "NO CUMPLE")
    cumple_cell.font = Font(name="Calibri", bold=True, color=COLOR_OK if calc.cumple_ric else COLOR_FAIL, size=11)
    ws.cell(row=r, column=1, value="Resultado")
    _style_row(ws.cell(row=r, column=1))
    r += 1

    # Footer
    r += 1
    ws.cell(row=r, column=1, value="Generado por RIC Conductor.calc — RIC")
    ws.cell(row=r, column=1).font = Font(name="Calibri", color="6E7681", size=8, italic=True)

    # Guardar en buffer
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)

    safe_name = sanitize_filename(calc.name or str(calc.id), fallback="calculo", max_length=60)
    filename = f"calculo_RIC_{safe_name}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )
